"""智能导入服务 - 文件抽取、Chunk 切片、LLM 解析、质量评分、自动入库/人工复核

核心流程：
  文件上传 → 创建 ImportJob/BackgroundJob → Worker 异步执行
  → 文本抽取 → Chunk 切片 → LLM 结构化解析 → Pydantic/程序校验
  → 高置信度自动入库 → 低置信度进入人工复核
"""

import hashlib
import json
import logging
import re
import time
from datetime import datetime, timezone
from decimal import Decimal

import httpx
from pydantic import ValidationError
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.storage import compute_file_hash, save_upload_file
from app.models.background_job import BackgroundJob
from app.models.bank_word import BankWordExclusion, BankWordFrequency
from app.models.import_chunk import ImportChunk
from app.models.import_job import ImportJob
from app.models.import_parsed_question import ImportParsedQuestion
from app.models.import_review_item import ImportReviewItem
from app.models.llm_parse_cache import LlmParseCache
from app.models.question import Question
from app.models.question_bank import QuestionBank
from app.schemas.llm_parse import LlmParseResult, ParsedQuestion
from app.services.ai_service import call_ai_api
from app.services.import_service import build_bank_word_frequencies
from app.services.job_service import (
    JOB_TYPE_QUESTION_IMPORT_LLM,
    JOB_TYPE_QUESTION_IMPORT_LLM_REPARSE,
    complete_job,
    fail_job,
    heartbeat_job,
    requeue_job,
)

logger = logging.getLogger(__name__)

# ─── 常量 ──────────────────────────────────────

PROMPT_VERSION = "v2-scenario-content"
CHUNK_MAX_CHARS = 12000
CHUNK_MIN_CHARS = 500

# ─── PR-2: 重试 / 降级常量 ─────────────────────
# L1 整 chunk 重试：首次调用 + L1_MAX_RETRIES 次重试 = 共 (L1_MAX_RETRIES + 1) 次。
# 重试前固定 sleep L1_RETRY_BACKOFF_SECONDS。
# RETRY_BASE_SECONDS / RETRY_CAP_SECONDS 为未来扩展"多次重试 + 指数退避"预留，
# PR-2 暂不消费。
L1_MAX_RETRIES = 1
L1_RETRY_BACKOFF_SECONDS = 2.0
RETRY_BASE_SECONDS = 2.0
RETRY_CAP_SECONDS = 10.0

# L2 单题降级：每段单独 LLM 调用使用更短的 timeout，
# 因为单题输入文本远小于 chunk 级别。
L2_PER_QUESTION_TIMEOUT = 60.0

# 单 chunk 总耗时上限（含 L1 + L2）。超出后剩余单题写入
# per_question_failures.stage="L2_fallback_budget_exceeded" 后跳出 L2 循环。
# 480s 约 2.7 个 worker lease 周期（DEFAULT_JOB_LEASE_SECONDS=180s），
# 配合 L2 循环内的 heartbeat_job 续约。
CHUNK_TOTAL_BUDGET_SECONDS = 480.0

# L2 循环内每 N 段调一次 heartbeat 续约 lease。
HEARTBEAT_EVERY_N_SEGMENTS = 3

# L1 重试候选异常类型。
# - httpx.TimeoutException 是 httpx.HTTPError 的子类；显式列出仅为可读性。
# - ValueError("AI API 错误 (5xx)") 走 _is_retryable_value_error 路径。
RETRYABLE_HTTP_EXC = (httpx.TimeoutException, httpx.HTTPError)

# 题号分割模式（优先级从高到低）
QUESTION_SPLIT_PATTERNS = [
    re.compile(r'(?:^|\n)\s*Question\s+#?\d+', re.IGNORECASE),
    re.compile(r'(?:^|\n)\s*QUESTION\s*[:#]\s*\d+', re.IGNORECASE),
    re.compile(r'(?:^|\n)\s*NEW\s+QUESTION\s+\d+', re.IGNORECASE),
    re.compile(r'(?:^|\n)\s*NO\.\s*\d+', re.IGNORECASE),
]

# 答案键检测模式
ANSWER_KEY_PATTERN = re.compile(
    r'(?:Answer\s*Key|Answers?\s*:|ANSWER\s*KEY|正确答案)[:\s]*\n([\s\S]+?)$',
    re.IGNORECASE,
)

# 答案键条目解析
ANSWER_ENTRY_PATTERN = re.compile(
    r'(\d{1,4})[.\s)]+\s*([A-Ea-e](?:\s*[,，]\s*[A-Ea-e])*|True|False)',
    re.IGNORECASE,
)

# 噪声关键词（页眉页脚常见内容）
NOISE_PATTERNS = [
    re.compile(r'CIPT\s+(?:Exam|Dumps|Questions|Practice)', re.IGNORECASE),
    re.compile(r'Page\s+\d+\s*(?:of|/)\s*\d+', re.IGNORECASE),
    re.compile(r'Passing\s+Score.*Time\s+Limit', re.IGNORECASE),
    re.compile(r'IAPP\s+Certified\s+Information', re.IGNORECASE),
    re.compile(r'www\.\S+\.(com|net|org)', re.IGNORECASE),
    re.compile(r'ExamQuestions\s+v\d', re.IGNORECASE),
    re.compile(r'by\s+Willow', re.IGNORECASE),
    re.compile(r'File\s+Version\s+\d', re.IGNORECASE),
]


# ─── 题干合成与去重签名 ─────────────────────────


def build_full_question_content(scenario_text: str | None, content: str | None) -> str:
    """合成正式题库题干：场景/背景材料在前，最后一问在后。

    ImportParsedQuestion 继续保留原始 scenario_text 和 content；正式 Question.content
    统一通过此 helper 生成，避免自动入库、人工复核、reparse、历史回填规则漂移。
    """
    scenario = (scenario_text or "").strip()
    stem = (content or "").strip()
    if scenario and stem:
        if _normalize_content_for_match(stem).startswith(_normalize_content_for_match(scenario)):
            return stem
        return f"{scenario}\n\n{stem}"
    return scenario or stem


def _normalize_content_for_match(text: str | None) -> str:
    """用于安全匹配的题干规范化：忽略首尾和连续空白差异。"""
    return re.sub(r"\s+", " ", (text or "").strip())


def _content_equivalent_for_backfill(current_content: str | None, parsed_content: str | None) -> bool:
    """历史回填保护条件：正式题干仍等价于解析出的短题干时才允许覆盖。"""
    return _normalize_content_for_match(current_content) == _normalize_content_for_match(parsed_content)


def _question_signature(
    question_type: str, content: str, options: list, correct_answer: list
) -> tuple:
    """生成题目唯一签名，用于去重（复用旧版逻辑，增强 options/answer 排序归一化）。"""

    def _option_label(opt: dict) -> str:
        return str(opt.get("label", opt.get("key", "")) or "").strip().upper()

    normalized_options = (
        json.dumps(
            sorted(
                [
                    {
                        "label": _option_label(opt),
                        "text": opt.get("text", ""),
                    }
                    for opt in options
                ],
                key=lambda o: o.get("label", ""),
            ),
            sort_keys=True, ensure_ascii=False, separators=(",", ":"),
        )
        if isinstance(options, list)
        else str(options)
    )
    normalized_answer = (
        ",".join(sorted(a.strip().upper() for a in correct_answer))
        if correct_answer
        else ""
    )
    return (
        question_type,
        (content or "").strip(),
        normalized_options,
        normalized_answer,
    )


def _is_retryable_value_error(exc: ValueError) -> bool:
    """判断 call_ai_api 抛出的 ValueError 是否可重试。

    仅服务端 5xx 视为瞬时；4xx / "API Key 未配置" 等不重试。
    依赖 ai_service.call_ai_api 现有的错误消息格式："AI API 错误 (5xx): ..."。
    """
    return "AI API 错误 (5" in str(exc)


def _normalize_qno(qno: str | None) -> str | None:
    """归一化题号：去前后空白 + 去 '#' 前缀 + 再去前后空白；空串返回 None。

    用于 imported_qnos 集合构建与 _save_parsed_question 入口比对，
    处理 LLM 输出 ``"#222"`` / ``" #222 "`` / ``"222"`` 等格式漂移。
    保持 ``str`` 而非 ``int`` —— 通用性考虑（如 ``"5a"`` / ``"5-1"`` 等非纯数字题号）。
    """
    if qno is None:
        return None
    cleaned = qno.strip().lstrip("#").strip()
    return cleaned or None


def _qno_sort_key(qno: str) -> tuple[int, int, str]:
    """题号排序键：纯数字按 int 升序（数字桶在前），非数字按字典序兜底。

    返回三元组以保证类型一致（避免 (0, int) vs (1, str) 直接比较 TypeError）。
    """
    if qno.isdigit():
        return (0, int(qno), "")
    return (1, 0, qno)


# ─── 创建导入任务 ──────────────────────────────────


def create_smart_import_job(
    db: Session,
    bank_id: int,
    file_bytes: bytes,
    filename: str,
    user_id: int,
    auto_import: bool = True,
    use_llm_cache: bool = True,
    force: bool = False,
) -> dict:
    """创建 ImportJob + BackgroundJob，返回导入任务信息"""
    bank = db.get(QuestionBank, bank_id)
    if not bank:
        return {"error": "题库不存在"}

    # 校验文件大小
    if len(file_bytes) > settings.upload_max_size_bytes:
        return {"error": f"文件大小超过限制（最大 {settings.MAX_UPLOAD_SIZE_MB}MB）"}

    # 确定文件类型
    filename_lower = (filename or "").lower()
    if filename_lower.endswith(".pdf"):
        file_type = "pdf"
    elif filename_lower.endswith(".xlsx"):
        file_type = "xlsx"
    elif filename_lower.endswith(".docx"):
        file_type = "docx"
    else:
        return {"error": f"不支持的文件格式: {filename}"}

    # 保存文件（先保存以获取 file_hash）
    file_path, file_hash = save_upload_file(file_bytes, filename)

    # 同文件重复导入仅记录溯源信息，不再硬阻断。
    # 去重边界在题目级：run_smart_import 会把题库现有 Question 签名加入 seen_signatures，
    # _save_parsed_question 命中后写入 duplicate/skipped 审计记录，避免重复写正式题库。
    duplicate_file_job = (
        db.query(ImportJob)
        .filter_by(bank_id=bank_id, file_hash=file_hash)
        .order_by(ImportJob.id.desc())
        .first()
    )
    config_json = {
        "auto_import": auto_import,
        "use_llm_cache": use_llm_cache,
    }
    if duplicate_file_job:
        config_json = {
            **config_json,
            "duplicate_file_of": duplicate_file_job.id,
            "duplicate_file_status": duplicate_file_job.status,
        }

    # 创建 ImportJob
    import_job = ImportJob(
        bank_id=bank_id,
        file_name=filename,
        file_path=file_path,
        file_hash=file_hash,
        file_type=file_type,
        status="pending",
        config_json=config_json,
        created_by=user_id,
    )
    db.add(import_job)
    db.flush()

    # 创建 BackgroundJob
    scope_key = f"import_llm:{import_job.id}"
    payload = {
        "import_job_id": import_job.id,
        "bank_id": bank_id,
        "file_path": file_path,
        "file_name": filename,
        "file_type": file_type,
        "auto_import": auto_import,
        "use_llm_cache": use_llm_cache,
    }
    bg_job = BackgroundJob(
        job_type=JOB_TYPE_QUESTION_IMPORT_LLM,
        scope_key=scope_key,
        active_scope_key=scope_key,
        payload_json=json.dumps(payload, ensure_ascii=False),
        status="queued",
        progress_total=0,
        status_message="等待后台 worker 执行智能导入",
        created_by=user_id,
    )
    db.add(bg_job)
    db.flush()

    import_job.background_job_id = bg_job.id
    db.commit()

    return {
        "import_job_id": import_job.id,
        "background_job_id": bg_job.id,
        "status": "pending",
    }


# ─── 主编排器 ──────────────────────────────────


def run_smart_import(db: Session, background_job: BackgroundJob) -> None:
    """智能导入主编排器，由 Worker 调用"""
    payload = _deserialize_payload(background_job)
    import_job_id = payload.get("import_job_id")
    if not import_job_id:
        raise ValueError("payload 缺少 import_job_id")

    import_job = db.get(ImportJob, import_job_id)
    if not import_job:
        raise ValueError(f"ImportJob {import_job_id} 不存在")

    file_path = payload.get("file_path", "")
    file_type = payload.get("file_type", "")
    auto_import = payload.get("auto_import", True)
    use_llm_cache = payload.get("use_llm_cache", True)

    # 阶段1：文本抽取
    _update_import_job_status(db, import_job, "extracting")
    try:
        pages = _extract_pages_from_file(file_path, file_type)
    except Exception as exc:
        _fail_import_job(db, import_job, f"文件抽取失败: {exc}")
        return

    if not pages:
        _fail_import_job(db, import_job, "文件抽取结果为空")
        return

    import_job.total_pages = len(pages)
    db.commit()

    # 阶段2：文本规范化与 Chunk 切片
    _update_import_job_status(db, import_job, "chunking")
    full_text = "\n".join(p["text"] for p in pages)
    normalized_text = _normalize_text(full_text)

    # 检测并提取答案键
    answer_key = _extract_answer_key(normalized_text)
    answer_key_text = ""
    if answer_key:
        # 移除答案键部分以避免重复解析
        normalized_text = ANSWER_KEY_PATTERN.sub("", normalized_text).strip()
        answer_key_text = _format_answer_key(answer_key)

    chunks_data = _split_into_chunks(pages, normalized_text, answer_key_text)

    if not chunks_data:
        _fail_import_job(db, import_job, "文本切片结果为空，无法继续解析")
        return

    # 保存 chunks 到数据库
    for chunk_data in chunks_data:
        chunk_hash = hashlib.sha256(chunk_data["chunk_text"].encode("utf-8")).hexdigest()
        chunk = ImportChunk(
            import_job_id=import_job.id,
            chunk_no=chunk_data["chunk_no"],
            start_page=chunk_data.get("start_page"),
            end_page=chunk_data.get("end_page"),
            chunk_text=chunk_data["chunk_text"],
            normalized_text=chunk_data.get("normalized_text", ""),
            chunk_hash=chunk_hash,
            status="pending",
        )
        db.add(chunk)
    db.flush()

    import_job.total_chunks = len(chunks_data)

    # PR-4: 计算并存储 expected_qnos —— reconciliation 的"应入库题号"基线。
    # 在切完 chunk 后一次性写入；reparse 路径不重写该字段（保证可重复对账）。
    expected_qnos: set[str] = set()
    for chunk_data in chunks_data:
        for seg in _split_by_single_question(chunk_data["chunk_text"]):
            norm = _normalize_qno(seg.get("source_question_no"))
            if norm is not None:
                expected_qnos.add(norm)

    # 将 answer_key_text 与 expected_qnos 一并持久化到 config_json
    config = import_job.config_json or {}
    if answer_key_text:
        config["answer_key_text"] = answer_key_text
    config["expected_qnos"] = sorted(expected_qnos, key=_qno_sort_key)
    # 重新赋值触发 SQLAlchemy JSONB dirty 检测（与 PR-3 等位姿势一致）
    import_job.config_json = config

    db.commit()

    # 阶段3：LLM 解析
    _update_import_job_status(db, import_job, "parsing")

    # 构建 seen_signatures：Bank 已有题目 + 本 Job 已入库题目
    seen_signatures = set()
    existing_questions = db.query(Question).filter_by(bank_id=import_job.bank_id).all()
    for eq in existing_questions:
        eq_options = eq.options if isinstance(eq.options, list) else json.loads(eq.options or "[]")
        eq_answer = [a.strip() for a in (eq.correct_answer or "").split(",")] if eq.correct_answer else []
        seen_signatures.add(_question_signature(eq.question_type, eq.content, eq_options, eq_answer))

    chunks = (
        db.query(ImportChunk)
        .filter_by(import_job_id=import_job.id)
        .order_by(ImportChunk.chunk_no.asc())
        .all()
    )

    # 更新 BackgroundJob progress_total 为实际 chunk 数
    bg_job = db.get(BackgroundJob, background_job.id)
    if bg_job and bg_job.progress_total == 0:
        bg_job.progress_total = len(chunks)
        db.commit()

    for index, chunk in enumerate(chunks, start=1):
        # 刷新 import_job 引用
        import_job = db.get(ImportJob, import_job_id)
        bg_job = db.get(BackgroundJob, background_job.id)

        try:
            _process_chunk(
                db=db,
                chunk=chunk,
                import_job=import_job,
                auto_import=auto_import,
                use_llm_cache=use_llm_cache,
                seen_signatures=seen_signatures,
                bg_job=bg_job,
            )
        except Exception as exc:
            logger.error("Chunk %d 解析失败: %s", chunk.chunk_no, exc)
            chunk.status = "failed"
            chunk.issues_json = {"error": str(exc)}
            import_job.failed_chunks = (import_job.failed_chunks or 0) + 1
            db.commit()
            # 单个 chunk 失败不中断整个导入

        # 心跳更新进度
        heartbeat_job(
            db,
            bg_job,
            success_increment=1,
            status_message=f"解析 Chunk {index}/{len(chunks)}",
        )

    # 阶段4：验证与汇总
    _update_import_job_status(db, import_job, "validating")
    import_job = db.get(ImportJob, import_job_id)
    _finalize_import(db, import_job)


def _process_chunk(
    db: Session,
    chunk: ImportChunk,
    import_job: ImportJob,
    auto_import: bool,
    use_llm_cache: bool,
    seen_signatures: set | None = None,
    bg_job: BackgroundJob | None = None,
    imported_qnos: set[str] | None = None,
) -> None:
    """处理单个 chunk：缓存 → L1 整 chunk LLM → L1 重试 → L2 单题降级 → 入库。

    状态机（chunk.status）：
        pending → parsing → {parsed_cached | parsed | parsed_retry |
                              parsed_fallback | parsed_partial | failed |
                              llm_failed | parse_failed}

    PR-2 新增取值（无 schema 迁移成本，String(32) 无 enum 约束）：
        parsed_retry      : L1 重试 1 次后成功
        parsed_fallback   : L1 失败 → L2 单题降级**全部成功**
        parsed_partial    : L2 单题降级**部分失败**（含 per_question_failures）

    chunk.issues_json 在 PR-2 后的 schema：
        {
          "chunk_issues": [...],         # LLM 返回的 chunk 级 issues（可能缺失）
          "retry_count": int,            # L1 实际重试次数（0 或 L1_MAX_RETRIES）
          "fallback_used": bool,         # L2 是否触发
          "per_question_failures": [
            {"source_question_no": str|None, "stage": str, "error": str}
          ],
          "fallback_meta": {...}         # 仅 fallback_used=True 时出现
        }

    LlmParseCache 写入策略：
        - 一次性成功 / L1 重试后成功 → 写缓存（PROMPT_VERSION + chunk_hash 为键）。
        - L2 单题降级（无论成功与否） → **不写缓存**。理由：缓存键按整 chunk hash，
          单题响应拼装出的 chunk 级响应不能代表真实 LLM 的 chunk 级一致性。
        - 任何 hard-fail（4xx / API Key 未配置 / parse_failed / L2 全部失败） → 不写。

    异常处理：
        - 不可重试错误（4xx ValueError / API Key 未配置 / Pydantic ValidationError /
          json.JSONDecodeError）：写 chunk.status="llm_failed"/"parse_failed" + raise，
          由 run_smart_import 外层 except 兜底标 failed 并 failed_chunks++。
        - 可重试错误（httpx.TimeoutException / 5xx ValueError / httpx.HTTPError）：
          走 L2 fallback；如 L2 全部失败 → status="failed", failed_chunks++,
          **不再 raise**（已显式记账）。
    """
    chunk.status = "parsing"
    db.commit()

    chunk_text = chunk.chunk_text
    config = import_job.config_json or {}
    answer_key_text = config.get("answer_key_text", "")
    chunk_started_at = time.monotonic()

    # 1) 查缓存（仅一次，L1 重试不会再触发缓存查找）
    cache_key = _build_cache_key(chunk.chunk_hash)
    cached = _lookup_llm_cache(db, cache_key) if use_llm_cache else None

    if cached:
        _process_chunk_cached(
            db=db,
            chunk=chunk,
            cached=cached,
            import_job=import_job,
            chunk_text=chunk_text,
            auto_import=auto_import,
            seen_signatures=seen_signatures,
            imported_qnos=imported_qnos,
            bg_job=bg_job,
        )
        return

    # 2) 构建 prompt + 持久化 llm_request_json（便于事后审计）
    messages = _build_llm_prompt(chunk_text, answer_key_text)
    chunk.llm_request_json = {"messages": messages}
    db.commit()

    # 3) L1 整 chunk 调用（含最多 1 次重试）
    retry_count = 0
    fallback_used = False
    fallback_reason: str | None = None
    per_question_failures: list[dict] = []
    merged_questions: list[dict] = []
    response_text: str | None = None
    extra_chunk_issues: list[dict] = []

    try:
        response_text, retry_count = _call_llm_with_l1_retry(messages, db, timeout=120.0)
    except RETRYABLE_HTTP_EXC:
        # httpx.TimeoutException / httpx.HTTPError：L1 用尽 → 进 L2
        fallback_used = True
        fallback_reason = "l1_retry_exhausted"
        retry_count = L1_MAX_RETRIES
    except ValueError as exc:
        if _is_retryable_value_error(exc):
            # 5xx ValueError：L1 用尽 → 进 L2
            fallback_used = True
            fallback_reason = "l1_retry_exhausted"
            retry_count = L1_MAX_RETRIES
        else:
            # 4xx / API Key 缺失：硬失败，不进 L2
            chunk.status = "llm_failed"
            chunk.issues_json = {
                **(chunk.issues_json or {}),
                "error": f"LLM 调用失败: {exc}",
                "retry_count": 0,
                "fallback_used": False,
                "per_question_failures": [],
            }
            db.commit()
            raise

    # 4) 如 L1 已决定降级，先执行 L2 单题降级。
    if fallback_used:
        merged_questions, per_question_failures = _run_per_question_fallback(
            chunk_text=chunk_text,
            answer_key_text=answer_key_text,
            db=db,
            bg_job=bg_job,
            chunk=chunk,
            started_at=chunk_started_at,
        )
        response_text = _build_fallback_response_text(merged_questions, per_question_failures)

    # 5) 解析 LLM 响应（L1 路径可能失败；L2 路径自构造，保证可解析）
    try:
        llm_result = _parse_llm_response(response_text or "")
    except Exception as exc:
        chunk.status = "parse_failed"
        chunk.issues_json = {
            **(chunk.issues_json or {}),
            "error": f"LLM 响应解析失败: {exc}",
            "raw_response": (response_text or "")[:500],
            "retry_count": retry_count,
            "fallback_used": fallback_used,
            "per_question_failures": per_question_failures,
        }
        chunk.llm_response_json = {"raw": (response_text or "")[:2000]}
        db.commit()
        raise

    # 5.1) L1 完整性检查：合法 JSON 但题量明显少于 chunk 内可检测题号段时，
    #      视为 chunk 级 LLM 漏题，改走 L2 单题降级。此时尚未保存任何题目，
    #      可安全丢弃不完整 L1 结果，避免重复写入。
    if not fallback_used:
        expected_segments = _split_by_single_question(chunk_text)
        expected_count = len(expected_segments)
        actual_count = len(llm_result.questions)
        if expected_count > 0 and actual_count < expected_count:
            fallback_used = True
            fallback_reason = "l1_incomplete_response"
            extra_chunk_issues.append({
                "code": "L1_INCOMPLETE_RESPONSE",
                "severity": "MEDIUM",
                "detail": (
                    f"L1 returned {actual_count} questions, "
                    f"but chunk contains {expected_count} detectable question segments; "
                    "switched to L2 per-question fallback"
                ),
                "expected_segments": expected_count,
                "actual_questions": actual_count,
            })
            logger.warning(
                "[smart_import] chunk_no=%s L1 incomplete response: expected_segments=%d actual_questions=%d; entering L2 fallback",
                chunk.chunk_no, expected_count, actual_count,
            )
            merged_questions, per_question_failures = _run_per_question_fallback(
                chunk_text=chunk_text,
                answer_key_text=answer_key_text,
                db=db,
                bg_job=bg_job,
                chunk=chunk,
                started_at=chunk_started_at,
            )
            response_text = _build_fallback_response_text(merged_questions, per_question_failures)
            llm_result = _parse_llm_response(response_text or "")

    # 6) 写 llm_response_json（结构化便于排查）
    try:
        chunk.llm_response_json = json.loads(response_text) if response_text else None
    except json.JSONDecodeError:
        chunk.llm_response_json = {"raw": (response_text or "")[:2000]}

    # 7) 计算最终 chunk.status
    if fallback_used and not merged_questions:
        # L2 全部失败 / 无法切段
        chunk.status = "failed"
        import_job.failed_chunks = (import_job.failed_chunks or 0) + 1
    elif fallback_used and per_question_failures:
        # L2 部分成功
        chunk.status = "parsed_partial"
        import_job.failed_chunks = (import_job.failed_chunks or 0) + 1
    elif fallback_used:
        # L2 全部成功
        chunk.status = "parsed_fallback"
    elif retry_count > 0:
        chunk.status = "parsed_retry"
    else:
        chunk.status = "parsed"

    # 8) 写 issues_json（PR-2 schema）
    issues_payload: dict = {
        **(chunk.issues_json or {}),
        "retry_count": retry_count,
        "fallback_used": fallback_used,
        "per_question_failures": per_question_failures,
    }
    if fallback_used:
        issues_payload["fallback_meta"] = {
            "total_segments": len(merged_questions) + len(per_question_failures),
            "succeeded": len(merged_questions),
            "failed": len(per_question_failures),
            "elapsed_seconds": round(time.monotonic() - chunk_started_at, 2),
        }
        if fallback_reason:
            issues_payload["fallback_meta"]["reason"] = fallback_reason
    chunk_issues = extra_chunk_issues + (llm_result.chunk_issues or [])
    if chunk_issues:
        issues_payload["chunk_issues"] = chunk_issues
    chunk.issues_json = issues_payload

    # 9) 缓存写入（仅 L1 路径成功 / L1 重试成功；L2 路径与失败路径都不写）
    if use_llm_cache and not fallback_used and chunk.status not in ("failed", "llm_failed", "parse_failed"):
        _store_llm_cache(
            db, cache_key, chunk.chunk_hash,
            request_json=chunk.llm_request_json,
            response_text=response_text or "",
        )

    db.commit()

    # 10) 入库（status="failed" 时 questions 必为空，但显式跳过更稳健）
    if chunk.status == "failed":
        return

    for parsed_q in llm_result.questions:
        _save_parsed_question(
            db=db,
            parsed_q=parsed_q,
            import_job=import_job,
            chunk=chunk,
            chunk_text=chunk_text,
            auto_import=auto_import,
            seen_signatures=seen_signatures,
            imported_qnos=imported_qnos,
        )


def _process_chunk_cached(
    db: Session,
    chunk: ImportChunk,
    cached: dict,
    import_job: ImportJob,
    chunk_text: str,
    auto_import: bool,
    seen_signatures: set | None,
    imported_qnos: set[str] | None = None,
    bg_job: BackgroundJob | None = None,
) -> None:
    """处理 LlmParseCache 命中的 chunk：直接复用历史响应，跳过 LLM 调用。"""
    response_text = cached.get("response_text", "")
    chunk.llm_request_json = cached.get("request_json")
    try:
        chunk.llm_response_json = json.loads(response_text) if response_text else None
    except json.JSONDecodeError:
        chunk.llm_response_json = {"raw": response_text[:2000]}

    try:
        llm_result = _parse_llm_response(response_text)
    except Exception as exc:
        chunk.status = "parse_failed"
        chunk.issues_json = {
            **(chunk.issues_json or {}),
            "error": f"缓存响应解析失败: {exc}",
        }
        db.commit()
        raise

    expected_segments = _split_by_single_question(chunk_text)
    expected_count = len(expected_segments)
    actual_count = len(llm_result.questions)
    if expected_count > 0 and actual_count < expected_count:
        logger.warning(
            "[smart_import] chunk_no=%s cached L1 response incomplete: expected_segments=%d actual_questions=%d; bypassing cache and entering L2 fallback",
            chunk.chunk_no, expected_count, actual_count,
        )
        config = import_job.config_json or {}
        answer_key_text = config.get("answer_key_text", "")
        started_at = time.monotonic()
        merged_questions, per_question_failures = _run_per_question_fallback(
            chunk_text=chunk_text,
            answer_key_text=answer_key_text,
            db=db,
            bg_job=bg_job,
            chunk=chunk,
            started_at=started_at,
        )
        response_text = _build_fallback_response_text(merged_questions, per_question_failures)
        llm_result = _parse_llm_response(response_text)
        try:
            chunk.llm_response_json = json.loads(response_text)
        except json.JSONDecodeError:
            chunk.llm_response_json = {"raw": response_text[:2000]}

        if not merged_questions:
            chunk.status = "failed"
            import_job.failed_chunks = (import_job.failed_chunks or 0) + 1
        elif per_question_failures:
            chunk.status = "parsed_partial"
            import_job.failed_chunks = (import_job.failed_chunks or 0) + 1
        else:
            chunk.status = "parsed_fallback"

        chunk.issues_json = {
            **(chunk.issues_json or {}),
            "retry_count": 0,
            "fallback_used": True,
            "per_question_failures": per_question_failures,
            "fallback_meta": {
                "total_segments": len(merged_questions) + len(per_question_failures),
                "succeeded": len(merged_questions),
                "failed": len(per_question_failures),
                "elapsed_seconds": round(time.monotonic() - started_at, 2),
                "reason": "l1_incomplete_response",
            },
            "chunk_issues": [{
                "code": "L1_INCOMPLETE_RESPONSE",
                "severity": "MEDIUM",
                "detail": (
                    f"Cached L1 response returned {actual_count} questions, "
                    f"but chunk contains {expected_count} detectable question segments; "
                    "switched to L2 per-question fallback"
                ),
                "expected_segments": expected_count,
                "actual_questions": actual_count,
            }] + (llm_result.chunk_issues or []),
        }
        db.commit()
        if chunk.status == "failed":
            return
    else:
        chunk.status = "parsed_cached"
        if llm_result.chunk_issues:
            chunk.issues_json = {
                **(chunk.issues_json or {}),
                "chunk_issues": llm_result.chunk_issues,
            }
        db.commit()

    for parsed_q in llm_result.questions:
        _save_parsed_question(
            db=db,
            parsed_q=parsed_q,
            import_job=import_job,
            chunk=chunk,
            chunk_text=chunk_text,
            auto_import=auto_import,
            seen_signatures=seen_signatures,
            imported_qnos=imported_qnos,
        )


def _build_fallback_response_text(
    merged_questions: list[dict],
    per_question_failures: list[dict],
) -> str:
    """把 L2 单题降级结果拼成 chunk 级伪响应，复用下游 Pydantic 解析逻辑。"""
    return json.dumps(
        {
            "questions": merged_questions,
            "chunk_issues": [],
            "_fallback_meta": {
                "total_segments": len(merged_questions) + len(per_question_failures),
                "succeeded": len(merged_questions),
                "failed": len(per_question_failures),
            },
        },
        ensure_ascii=False,
    )


def _call_llm_with_l1_retry(
    messages: list[dict],
    db: Session,
    *,
    timeout: float,
    max_retries: int = L1_MAX_RETRIES,
    backoff: float = L1_RETRY_BACKOFF_SECONDS,
) -> tuple[str, int]:
    """整 chunk 调用 LLM；遇可重试异常自动重试 max_retries 次。

    返回:
        (response_text, attempts_after_first)：attempts_after_first=0 表示首次成功，
        attempts_after_first=N 表示重试 N 次后成功。

    抛出:
        最后一次失败的异常（httpx.HTTPError / httpx.TimeoutException / 5xx ValueError）；
        非可重试的 ValueError（4xx / API Key 缺失）直接首次抛出，不重试。

    决策记录（详见 PR-2 design 文档 D.1）：
        当前 max_retries=L1_MAX_RETRIES=1，固定 sleep backoff（不指数）。
        指数退避常量 RETRY_BASE_SECONDS / RETRY_CAP_SECONDS 为未来扩展预留。
    """
    last_exc: Exception | None = None
    for attempt in range(max_retries + 1):
        try:
            text = call_ai_api(messages, db, scene="smart_import", timeout=timeout)
            return text, attempt
        except ValueError as exc:
            if not _is_retryable_value_error(exc):
                raise
            last_exc = exc
        except RETRYABLE_HTTP_EXC as exc:
            last_exc = exc
        if attempt < max_retries:
            logger.warning(
                "[smart_import] L1 retry attempt=%d/%d sleep=%ss reason=%s",
                attempt + 1, max_retries, backoff, type(last_exc).__name__,
            )
            time.sleep(backoff)
    # 用尽重试次数；last_exc 必非 None
    assert last_exc is not None
    raise last_exc


def _run_per_question_fallback(
    chunk_text: str,
    answer_key_text: str,
    db: Session,
    bg_job: BackgroundJob | None,
    chunk: ImportChunk,
    started_at: float,
) -> tuple[list[dict], list[dict]]:
    """L2 单题降级。把 chunk_text 切分成"每题一段"，每段单独发起 LLM 调用。

    返回:
        (merged_questions, per_question_failures)
        - merged_questions: 成功解析的题目 dict 列表（与 ParsedQuestion.model_dump() 同构）
        - per_question_failures: [{"source_question_no", "stage", "error"}]

    时间预算 / heartbeat：
        - 总耗时上限 CHUNK_TOTAL_BUDGET_SECONDS 由 started_at 推算。
          超时把剩余段标 stage="L2_fallback_budget_exceeded" 并跳出循环。
        - 每 HEARTBEAT_EVERY_N_SEGMENTS 段调一次 heartbeat_job 续约 lease；
          heartbeat 失败本身不应中断 fallback。
    """
    segments = _split_by_single_question(chunk_text)
    merged: list[dict] = []
    failures: list[dict] = []

    if not segments:
        return merged, [{
            "source_question_no": None,
            "stage": "L2_fallback_skipped",
            "error": "no_question_markers",
        }]

    total_segments = len(segments)
    logger.warning(
        "[smart_import] chunk_no=%s entering L2 per-question fallback (segments=%d)",
        chunk.chunk_no, total_segments,
    )
    for idx, seg in enumerate(segments, start=1):
        # 总耗时预算 kill switch（含已耗费的 L1 时间）
        if time.monotonic() - started_at > CHUNK_TOTAL_BUDGET_SECONDS:
            logger.warning(
                "[smart_import] chunk_no=%s L2 budget %ss exceeded, dropped=%d",
                chunk.chunk_no, CHUNK_TOTAL_BUDGET_SECONDS, len(segments) - idx + 1,
            )
            for rem in segments[idx - 1:]:
                failures.append({
                    "source_question_no": rem.get("source_question_no"),
                    "stage": "L2_fallback_budget_exceeded",
                    "error": f"chunk total budget {CHUNK_TOTAL_BUDGET_SECONDS}s exceeded",
                })
            break

        msgs = _build_llm_prompt(seg["text"], answer_key_text)
        try:
            txt = call_ai_api(
                msgs, db, scene="smart_import", timeout=L2_PER_QUESTION_TIMEOUT,
            )
            parsed = _parse_llm_response(txt)
            for q in parsed.questions:
                merged.append(q.model_dump())
        except (
            ValueError,
            httpx.HTTPError,
            json.JSONDecodeError,
            ValidationError,
        ) as exc:
            failures.append({
                "source_question_no": seg.get("source_question_no"),
                "stage": "L2_fallback",
                "error": f"{type(exc).__name__}: {exc}",
            })

        # heartbeat 续约（每 N 段一次；失败不中断 fallback）
        if bg_job is not None and idx % HEARTBEAT_EVERY_N_SEGMENTS == 0:
            try:
                heartbeat_job(
                    db, bg_job,
                    success_increment=0,
                    status_message=(
                        f"Chunk {chunk.chunk_no} L2 fallback {idx}/{total_segments}"
                    ),
                )
            except Exception as exc:
                # heartbeat 失败不应中断 fallback；继续下一段
                logger.warning(
                    "[smart_import] chunk_no=%s heartbeat failed at segment %d/%d: %s",
                    chunk.chunk_no, idx, total_segments, exc,
                )

    return merged, failures


def _persist_duplicate_parsed_question(
    db: Session,
    parsed_q: ParsedQuestion,
    import_job: ImportJob,
    chunk: ImportChunk,
    *,
    reason: str,
) -> None:
    """统一处理 DUPLICATE 路径：把 ImportParsedQuestion 写为 ``review_status='duplicate'``,
    ``import_status='skipped'``，**不**写入 Question 表，**不**更新 ``seen_signatures``。

    PR-3 起两条 DUPLICATE 路径（题号去重 + 内容签名）走同一 helper，杜绝双写漂移。

    issues_json 规约：
        - 主 ``code`` 保持 ``"DUPLICATE"`` —— 前端 / serialize_parsed_question 零兼容代价。
        - ``details[0].reason`` ∈ {"qno", "content"} —— PR-4 reconciliation 的稳定 lookup key。
    """
    if reason not in ("qno", "content"):
        raise ValueError(f"_persist_duplicate_parsed_question: unsupported reason={reason!r}")

    correct_answer_list = parsed_q.correct_answer or []
    question_type = parsed_q.question_type
    if question_type == "unknown":
        question_type = "single"

    options_for_storage = [
        {"key": opt.label, "text": opt.text} for opt in parsed_q.options
    ]
    correct_answer_str = ",".join(correct_answer_list) if correct_answer_list else ""

    detail = (
        f"题号 {_normalize_qno(parsed_q.source_question_no)} 已入库（reparse 跳过）"
        if reason == "qno"
        else "与已有题目重复"
    )

    parsed_question = ImportParsedQuestion(
        import_job_id=import_job.id,
        chunk_id=chunk.id,
        source_question_no=parsed_q.source_question_no,
        question_type=question_type,
        scenario_text=parsed_q.scenario,
        content=parsed_q.content,
        options_json=options_for_storage,
        correct_answer=correct_answer_str.split(",") if correct_answer_str else [],
        explanation=parsed_q.explanation or None,
        references_json=parsed_q.references if parsed_q.references else None,
        llm_confidence=Decimal(str(round(parsed_q.confidence, 4))),
        final_confidence=Decimal("0"),
        issues_json={
            "issues": ["DUPLICATE"],
            "details": [{
                "code": "DUPLICATE",
                "severity": "LOW",
                "reason": reason,
                "detail": detail,
            }],
        },
        review_status="duplicate",
        import_status="skipped",
    )
    logger.info(
        "[smart_import] duplicate parsed question reason=%s source_qno=%s",
        reason, parsed_q.source_question_no,
    )
    db.add(parsed_question)
    import_job.parsed_questions = (import_job.parsed_questions or 0) + 1
    db.commit()


def _unusable_question_issues(parsed_q: ParsedQuestion) -> list[dict]:
    issues = []
    full_content = build_full_question_content(parsed_q.scenario, parsed_q.content)
    if not full_content or len(full_content.strip()) < 10:
        issues.append({"code": "STEM_MISSING", "severity": "HIGH", "detail": "缺少题干或题干过短"})

    if len(parsed_q.options or []) < 2:
        issues.append({"code": "OPTIONS_MISSING", "severity": "HIGH", "detail": "选项不足"})

    if not parsed_q.correct_answer:
        issues.append({"code": "ANSWER_MISSING", "severity": "HIGH", "detail": "缺少正确答案"})
    else:
        option_labels = {opt.label.upper() for opt in parsed_q.options or []}
        if any(ans.upper() not in option_labels for ans in parsed_q.correct_answer):
            issues.append({"code": "ANSWER_NOT_IN_OPTIONS", "severity": "HIGH", "detail": "正确答案不在选项中"})

    return issues


def _merge_issue_details(quality_issues: list[dict], usability_issues: list[dict]) -> list[dict]:
    merged = []
    seen = set()
    for issue in usability_issues + quality_issues:
        code = issue.get("code")
        if code in seen:
            continue
        seen.add(code)
        merged.append(issue)
    return merged


def _save_parsed_question(
    db: Session,
    parsed_q: ParsedQuestion,
    import_job: ImportJob,
    chunk: ImportChunk,
    chunk_text: str,
    auto_import: bool,
    seen_signatures: set | None = None,
    imported_qnos: set[str] | None = None,
) -> None:
    """保存单个解析题目，执行质量评分并决定自动入库或人工复核。

    去重优先级（自上而下）：
        1. **DUPLICATE_QNO**（PR-3 引入，仅 reparse 路径生效）：
           ``imported_qnos`` 不为 None 时，若归一化后题号命中集合 → 走 DUPLICATE 路径；
           不更新 ``seen_signatures`` / ``imported_qnos``。
        2. **DUPLICATE_CONTENT**（既有内容签名）：
           ``seen_signatures`` 命中 → 走 DUPLICATE 路径；不更新 ``seen_signatures``。
        3. 否则：执行质量评分和可用性判断 → 自动入库 / 自动跳过 / 创建 ReviewItem。

    初次导入路径（``run_smart_import``）传 ``imported_qnos=None``，DUPLICATE_QNO 永不触发。
    """
    # PR-3：题号去重（reparse 路径）—— 优先于内容签名
    if imported_qnos is not None:
        norm_qno = _normalize_qno(parsed_q.source_question_no)
        if norm_qno is not None and norm_qno in imported_qnos:
            _persist_duplicate_parsed_question(
                db, parsed_q, import_job, chunk, reason="qno",
            )
            return

    # 内容签名去重（保留现有路径，PR-3 仅给 issues_json 加 reason="content"）
    correct_answer_list = parsed_q.correct_answer or []
    question_type = parsed_q.question_type
    if question_type == "unknown":
        question_type = "single"

    options_for_sig = [{"label": opt.label, "text": opt.text} for opt in parsed_q.options]
    full_content_for_sig = build_full_question_content(parsed_q.scenario, parsed_q.content)
    sig = _question_signature(question_type, full_content_for_sig, options_for_sig, correct_answer_list)
    if seen_signatures is not None and sig in seen_signatures:
        _persist_duplicate_parsed_question(
            db, parsed_q, import_job, chunk, reason="content",
        )
        return

    if seen_signatures is not None:
        seen_signatures.add(sig)

    # 质量检查
    final_confidence, quality_issues = _quality_check(parsed_q, chunk_text)
    unusable_issues = _unusable_question_issues(parsed_q)
    issues = _merge_issue_details(quality_issues, unusable_issues)

    # 转换 options 格式：ParsedOption(label, text) -> dict(key, text)
    options_for_storage = [{"key": opt.label, "text": opt.text} for opt in parsed_q.options]

    # 转换 correct_answer 格式：list[str] -> str
    correct_answer_str = ",".join(correct_answer_list) if correct_answer_list else ""

    # 创建 ImportParsedQuestion
    parsed_question = ImportParsedQuestion(
        import_job_id=import_job.id,
        chunk_id=chunk.id,
        source_question_no=parsed_q.source_question_no,
        question_type=question_type,
        scenario_text=parsed_q.scenario,
        content=parsed_q.content,
        options_json=options_for_storage,
        correct_answer=correct_answer_str.split(",") if correct_answer_str else [],
        explanation=parsed_q.explanation or None,
        references_json=parsed_q.references if parsed_q.references else None,
        llm_confidence=Decimal(str(round(parsed_q.confidence, 4))),
        final_confidence=final_confidence,
        issues_json={"issues": [i["code"] for i in issues], "details": issues} if issues else None,
        review_status="pending",
        import_status="waiting",
    )
    db.add(parsed_question)
    db.flush()

    if unusable_issues:
        parsed_question.import_status = "skipped"
        parsed_question.review_status = "auto_skipped"
    elif auto_import:
        question = _write_question_to_bank(db, parsed_question, import_job.bank_id)
        if parsed_question.import_status == "skipped":
            # _write_question_to_bank 的双重保险命中重复题时，会保留当前解析记录为
            # duplicate/skipped，并且不关联已有正式题，避免被计入已导入题。
            pass
        else:
            parsed_question.import_status = "imported"
            parsed_question.review_status = "auto_accepted"
            parsed_question.imported_question_id = question.id if question else None
            import_job.imported_questions = (import_job.imported_questions or 0) + 1
    else:
        severity = "HIGH" if any(i.get("severity") == "HIGH" for i in issues) else "MEDIUM"
        review_type = issues[0]["code"] if issues else "LOW_CONFIDENCE"
        review_item = ImportReviewItem(
            import_job_id=import_job.id,
            parsed_question_id=parsed_question.id,
            review_type=review_type,
            severity=severity,
            before_json={
                "content": parsed_q.content,
                "options": options_for_storage,
                "correct_answer": parsed_q.correct_answer,
            },
            status="pending",
        )
        db.add(review_item)
        import_job.review_questions = (import_job.review_questions or 0) + 1

    import_job.parsed_questions = (import_job.parsed_questions or 0) + 1
    db.commit()


def _write_question_to_bank(
    db: Session,
    parsed_question: ImportParsedQuestion,
    bank_id: int,
) -> Question | None:
    """将解析结果写入 Question 表（双重保险：再次检查 Question 表去重）"""
    # 转换 options 和 correct_answer 格式
    options = parsed_question.options_json
    if isinstance(options, list):
        options = [
            {"key": opt.get("key", opt.get("label", "")), "text": opt.get("text", "")}
            for opt in options
        ]

    correct_answer = parsed_question.correct_answer
    if isinstance(correct_answer, list):
        correct_answer_list = correct_answer
        correct_answer_str = ",".join(correct_answer) if correct_answer else ""
    else:
        correct_answer_str = str(correct_answer) if correct_answer else ""
        correct_answer_list = [a.strip() for a in correct_answer_str.split(",")] if correct_answer_str else []

    full_content = build_full_question_content(
        parsed_question.scenario_text,
        parsed_question.content,
    )

    # 双重保险：检查 Question 表是否已存在相同签名
    sig = _question_signature(
        parsed_question.question_type or "single",
        full_content,
        options,
        correct_answer_list,
    )
    existing = (
        db.query(Question)
        .filter_by(bank_id=bank_id, content=full_content)
        .first()
    )
    if existing:
        # 比较完整签名
        existing_sig = _question_signature(
            existing.question_type,
            existing.content,
            existing.options if isinstance(existing.options, list) else json.loads(existing.options or "[]"),
            [a.strip() for a in (existing.correct_answer or "").split(",")] if existing.correct_answer else [],
        )
        if existing_sig == sig:
            logger.info("题库 %d 已存在相同题目 (id=%d)，跳过写入", bank_id, existing.id)
            parsed_question.import_status = "skipped"
            parsed_question.review_status = "duplicate"
            parsed_question.imported_question_id = None
            parsed_question.issues_json = {
                "issues": ["DUPLICATE"],
                "details": [{
                    "code": "DUPLICATE",
                    "severity": "LOW",
                    "reason": "content",
                    "detail": "与已有题目重复",
                }],
            }
            db.flush()
            return existing

    # 获取当前最大 order_index
    max_order = db.query(func.max(Question.order_index)).filter_by(bank_id=bank_id).scalar() or -1

    question = Question(
        bank_id=bank_id,
        question_type=parsed_question.question_type or "single",
        content=full_content,
        options=options,
        correct_answer=correct_answer_str,
        explanation=parsed_question.explanation,
        order_index=max_order + 1,
    )
    db.add(question)
    db.flush()

    # 回写 imported_question_id
    parsed_question.imported_question_id = question.id
    db.flush()

    return question


# ─── 重新解析 ──────────────────────────────────


def create_reparse_job(
    db: Session,
    import_job_id: int,
    chunk_id: int,
    user_id: int,
) -> dict:
    """为单个 chunk 创建重新解析的 BackgroundJob"""
    import_job = db.get(ImportJob, import_job_id)
    if not import_job:
        return {"error": "导入任务不存在"}

    chunk = db.get(ImportChunk, chunk_id)
    if not chunk or chunk.import_job_id != import_job_id:
        return {"error": "Chunk 不存在或不属于该导入任务"}

    # 检查是否已有进行中的 reparse 任务
    scope_key = f"import_reparse:{chunk_id}"
    existing = (
        db.query(BackgroundJob)
        .filter_by(active_scope_key=scope_key)
        .order_by(BackgroundJob.id.desc())
        .first()
    )
    if existing and existing.status in ("queued", "running"):
        return {
            "background_job_id": existing.id,
            "status": existing.status,
            "message": "该 chunk 已有重新解析任务正在执行",
        }

    payload = {
        "import_job_id": import_job_id,
        "chunk_id": chunk_id,
        "bank_id": import_job.bank_id,
    }
    bg_job = BackgroundJob(
        job_type=JOB_TYPE_QUESTION_IMPORT_LLM_REPARSE,
        scope_key=scope_key,
        active_scope_key=scope_key,
        payload_json=json.dumps(payload, ensure_ascii=False),
        status="queued",
        progress_total=1,
        status_message="等待后台 worker 执行重新解析",
        created_by=user_id,
    )
    db.add(bg_job)
    db.commit()

    return {
        "background_job_id": bg_job.id,
        "status": "queued",
        "message": "重新解析任务已创建",
    }


def run_reparse(db: Session, background_job: BackgroundJob) -> None:
    """重新解析单个 chunk"""
    payload = _deserialize_payload(background_job)
    import_job_id = payload.get("import_job_id")
    chunk_id = payload.get("chunk_id")
    bank_id = payload.get("bank_id")

    import_job = db.get(ImportJob, import_job_id)
    if not import_job:
        raise ValueError(f"ImportJob {import_job_id} 不存在")

    chunk = db.get(ImportChunk, chunk_id)
    if not chunk:
        raise ValueError(f"ImportChunk {chunk_id} 不存在")

    # 删除该 chunk 下未导入的 parsed_questions 及关联的 review_items
    old_parsed = (
        db.query(ImportParsedQuestion)
        .filter_by(chunk_id=chunk_id)
        .all()
    )
    for pq in old_parsed:
        if pq.import_status == "imported" and pq.imported_question_id:
            # 已导入的题目保留，不删除
            continue
        # 删除关联的 review_items
        db.query(ImportReviewItem).filter_by(parsed_question_id=pq.id).delete()
        # 更新 import_job 统计
        if pq.review_status == "pending":
            import_job.review_questions = max(0, (import_job.review_questions or 0) - 1)
        import_job.parsed_questions = max(0, (import_job.parsed_questions or 0) - 1)
        db.delete(pq)

    db.flush()

    # 重置 chunk 状态
    chunk.status = "pending"
    chunk.llm_request_json = None
    chunk.llm_response_json = None
    chunk.issues_json = None
    db.commit()

    # 重新处理该 chunk
    config = import_job.config_json or {}
    auto_import = config.get("auto_import", True)
    use_llm_cache = False  # 重新解析时跳过缓存

    # 构建已入库题目的 seen_signatures
    seen_signatures = set()
    existing_questions = db.query(Question).filter_by(bank_id=import_job.bank_id).all()
    for eq in existing_questions:
        eq_options = eq.options if isinstance(eq.options, list) else json.loads(eq.options or "[]")
        eq_answer = [a.strip() for a in (eq.correct_answer or "").split(",")] if eq.correct_answer else []
        seen_signatures.add(_question_signature(eq.question_type, eq.content, eq_options, eq_answer))

    # PR-3：构建本 ImportJob 已入库的题号集合（reparse 卫生）。
    # 来源 = ImportParsedQuestion 表（按 import_job_id == this_job AND import_status == 'imported'）。
    # 选用 ImportParsedQuestion 而非反查 Question 表的理由：
    #   - source_question_no 字段直接，无需 join；
    #   - 仅本 job 范围，避免与同 bank 的其他来源题（手工导入等）混淆；
    #   - 与 PR-4 reconciliation 的同表切片保持 schema 一致。
    imported_qnos: set[str] = set()
    for pq in (
        db.query(ImportParsedQuestion)
        .filter_by(import_job_id=import_job.id, import_status="imported")
        .all()
    ):
        normalized = _normalize_qno(pq.source_question_no)
        if normalized is not None:
            imported_qnos.add(normalized)

    _process_chunk(
        db=db,
        chunk=chunk,
        import_job=import_job,
        auto_import=auto_import,
        use_llm_cache=use_llm_cache,
        seen_signatures=seen_signatures,
        bg_job=background_job,
        imported_qnos=imported_qnos,
    )

    # 更新 import_job 状态
    if import_job.review_questions > 0:
        _update_import_job_status(db, import_job, "review_required")
    else:
        _update_import_job_status(db, import_job, "imported")

    # 更新题库统计
    _update_bank_stats(db, import_job.bank_id)

    # PR-4：reparse 后也刷新 reconciliation。
    # 注意：reparse **不**走 _finalize_import（避免覆盖 status 状态机），
    # 只复用 _compute_reconciliation 重新计算对账数据 —— 满足 PRD AC3。
    _write_reconciliation(db, import_job)


# ─── 复核操作 ──────────────────────────────────


def accept_review_item(
    db: Session,
    import_job_id: int,
    review_item_id: int,
    reviewer_id: int,
) -> dict:
    """接受复核项，按 LLM 解析结果原样写入 Question"""
    import_job = db.get(ImportJob, import_job_id)
    if not import_job:
        return {"error": "导入任务不存在"}

    review_item = db.get(ImportReviewItem, review_item_id)
    if not review_item or review_item.import_job_id != import_job_id:
        return {"error": "复核项不存在或不属于该导入任务"}

    if review_item.status != "pending":
        return {"error": f"复核项状态为 {review_item.status}，无法再次操作"}

    parsed_question = db.get(ImportParsedQuestion, review_item.parsed_question_id)
    if not parsed_question:
        return {"error": "关联的解析题目不存在"}

    # 写入 Question
    question = _write_question_to_bank(db, parsed_question, import_job.bank_id)

    # 更新状态
    parsed_question.import_status = "imported"
    parsed_question.review_status = "accepted"
    parsed_question.imported_question_id = question.id
    review_item.status = "accepted"
    review_item.reviewer_id = reviewer_id
    review_item.reviewed_at = datetime.now(timezone.utc)

    import_job.imported_questions = (import_job.imported_questions or 0) + 1
    import_job.review_questions = max(0, (import_job.review_questions or 0) - 1)
    db.commit()

    # 更新题库统计
    _update_bank_stats(db, import_job.bank_id)

    return {"question_id": question.id, "message": "题目已入库"}


def skip_review_item(
    db: Session,
    import_job_id: int,
    review_item_id: int,
    reviewer_id: int,
) -> dict:
    """跳过复核项，不入库"""
    import_job = db.get(ImportJob, import_job_id)
    if not import_job:
        return {"error": "导入任务不存在"}

    review_item = db.get(ImportReviewItem, review_item_id)
    if not review_item or review_item.import_job_id != import_job_id:
        return {"error": "复核项不存在或不属于该导入任务"}

    if review_item.status != "pending":
        return {"error": f"复核项状态为 {review_item.status}，无法再次操作"}

    parsed_question = db.get(ImportParsedQuestion, review_item.parsed_question_id)

    # 更新状态
    if parsed_question:
        parsed_question.import_status = "skipped"
        parsed_question.review_status = "skipped"
    review_item.status = "skipped"
    review_item.reviewer_id = reviewer_id
    review_item.reviewed_at = datetime.now(timezone.utc)

    import_job.review_questions = max(0, (import_job.review_questions or 0) - 1)
    db.commit()

    return {"message": "题目已跳过"}


# ─── 文件抽取 ──────────────────────────────────


def _extract_pages_from_file(file_path: str, file_type: str) -> list[dict]:
    """从文件中逐页/逐表抽取文本，返回 [{page_no, text}]"""
    if file_type == "pdf":
        return _extract_pdf_pages(file_path)
    elif file_type == "xlsx":
        return _extract_xlsx_pages(file_path)
    elif file_type == "docx":
        return _extract_docx_pages(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {file_type}")


def _extract_pdf_pages(file_path: str) -> list[dict]:
    """使用 pdfplumber 逐页抽取 PDF 文本"""
    import pdfplumber

    pages = []
    with pdfplumber.open(file_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            if text and text.strip():
                pages.append({"page_no": i, "text": _clean_text(text)})
    return pages


def _extract_xlsx_pages(file_path: str) -> list[dict]:
    """使用 openpyxl 逐 sheet 抽取 XLSX 文本"""
    from openpyxl import load_workbook

    pages = []
    wb = load_workbook(file_path, read_only=True)
    for sheet_index, ws in enumerate(wb.worksheets, start=1):
        lines = []
        for row in ws.iter_rows(values_only=True):
            line = " ".join(str(c) for c in row if c is not None)
            if line.strip():
                lines.append(line)
        if lines:
            pages.append({
                "page_no": sheet_index,
                "text": "\n".join(lines),
            })
    wb.close()
    return pages


def _extract_docx_pages(file_path: str) -> list[dict]:
    """使用 python-docx 抽取 DOCX 文本（视为单页）"""
    from docx import Document

    doc = Document(file_path)
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if text:
        return [{"page_no": 1, "text": text}]
    return []


def _clean_text(text: str) -> str:
    """清理 PDF 提取文本中的控制字符和常见 ligature 问题"""
    # 复用 import_service 中的清洗逻辑
    from app.services.import_service import _clean_text as _import_clean
    return _import_clean(text)


# ─── 文本处理 ──────────────────────────────────


def _normalize_text(text: str) -> str:
    """规范化文本：合并连续空白、去除控制字符"""
    # 去除控制字符
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    # 合并连续空白行为单个换行
    text = re.sub(r'\n{3,}', '\n\n', text)
    # 合并行内连续空格
    text = re.sub(r'[^\S\n]{2,}', ' ', text)
    return text.strip()


def _extract_answer_key(text: str) -> dict[int, str]:
    """从文本末尾提取答案键（如 Answer Key 部分）"""
    match = ANSWER_KEY_PATTERN.search(text)
    if not match:
        return {}

    key_text = match.group(1)
    answer_key = {}
    for m in ANSWER_ENTRY_PATTERN.finditer(key_text):
        q_num = int(m.group(1))
        answer = m.group(2).strip().upper().replace("，", ",")
        answer_key[q_num] = answer

    return answer_key if len(answer_key) >= 3 else {}  # 至少3条才算答案键


def _format_answer_key(answer_key: dict[int, str]) -> str:
    """将答案键格式化为文本供 LLM 参考"""
    lines = []
    for q_num in sorted(answer_key.keys()):
        lines.append(f"{q_num}. {answer_key[q_num]}")
    return "\n".join(lines)


def _split_into_chunks(
    pages: list[dict],
    normalized_text: str,
    answer_key_text: str,
) -> list[dict]:
    """将文本按题号切分成 chunks，每个 chunk 包含 1-5 道题

    返回: [{chunk_no, start_page, end_page, chunk_text, normalized_text}]
    答案参考表不再拼入 chunk_text，通过 _build_llm_prompt() 的 system prompt 传递
    """
    # 尝试按题号模式分割
    segments = _split_by_question_markers(normalized_text)

    if not segments:
        # 没有找到题号模式，按字符数粗切
        segments = _split_by_char_count(normalized_text)

    chunks = []
    for i, segment in enumerate(segments, start=1):
        chunks.append({
            "chunk_no": i,
            "start_page": segment.get("start_page"),
            "end_page": segment.get("end_page"),
            "chunk_text": segment["text"],
            "normalized_text": segment["text"],
        })

    return chunks


def _looks_like_leading_reading_material(text: str) -> bool:
    """保守判断第一个题号前文本是否应归入后续第一题。

    只接收明显像场景/阅读材料的前导文本；拒绝答案段、选项段、页眉广告等噪声，
    以降低把上一题解析或页眉误拼到下一题的风险。
    """
    leading = (text or "").strip()
    if not leading:
        return False

    if len(leading) < 80 and not re.search(r"\b(SCENARIO|CASE\s+STUDY)\b", leading, re.IGNORECASE):
        return False

    reject_patterns = [
        r"\b(Correct\s+Answer|Answer\s*:|Answer\s+Key|Explanation\s*:|Reference\s*:)\b",
        r"(?:^|\n)\s*[A-H][.)]\s+\S+",
        r"\b(?:Page\s+\d+\s*(?:of|/)|www\.\S+|ExamQuestions\s+v\d|Passing\s+Score)\b",
    ]
    if any(re.search(pattern, leading, re.IGNORECASE) for pattern in reject_patterns):
        return False

    has_scenario_marker = re.search(r"\b(SCENARIO|CASE\s+STUDY)\b", leading, re.IGNORECASE)
    paragraph_count = len([p for p in re.split(r"\n\s*\n", leading) if p.strip()])
    sentence_count = len(re.findall(r"[.!?](?:\s|$)", leading))
    word_count = len(re.findall(r"\w+", leading))

    return bool(has_scenario_marker or paragraph_count >= 2 or sentence_count >= 3 or word_count >= 80)


def _split_by_question_markers(text: str) -> list[dict]:
    """使用题号模式将文本分割为片段"""
    # 找到所有分割点
    best_pattern = None
    best_count = 0

    for pattern in QUESTION_SPLIT_PATTERNS:
        matches = list(pattern.finditer(text))
        if len(matches) > best_count:
            best_count = len(matches)
            best_pattern = pattern

    if best_count == 0 or not best_pattern:
        return []

    matches = list(best_pattern.finditer(text))
    segments = []
    leading_text = text[:matches[0].start()]
    attach_leading_to_first = _looks_like_leading_reading_material(leading_text)

    for i, match in enumerate(matches):
        start = 0 if i == 0 and attach_leading_to_first else match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        segment_text = text[start:end].strip()

        if segment_text:
            segments.append({"text": segment_text, "start_page": None, "end_page": None})

    # 将过小的片段合并
    merged = []
    buffer_text = ""
    for seg in segments:
        if len(buffer_text) + len(seg["text"]) > CHUNK_MAX_CHARS and buffer_text:
            merged.append({"text": buffer_text, "start_page": None, "end_page": None})
            buffer_text = seg["text"]
        else:
            buffer_text = buffer_text + "\n\n" + seg["text"] if buffer_text else seg["text"]

    if buffer_text:
        merged.append({"text": buffer_text, "start_page": None, "end_page": None})

    return merged if merged else segments


def _split_by_single_question(text: str) -> list[dict]:
    """与 _split_by_question_markers 类似，但每段恰含 1 道题（不合并短片段）。

    用途：L2 单题降级路径下，把整 chunk 文本切成"每题一段"，每段独立调 LLM。

    返回:
        [{"text": str, "source_question_no": str | None}, ...]
        若文本无任何题号正则命中，返回空列表。
    """
    best_pattern = None
    best_count = 0
    for pattern in QUESTION_SPLIT_PATTERNS:
        matches = list(pattern.finditer(text))
        if len(matches) > best_count:
            best_count = len(matches)
            best_pattern = pattern

    if best_count == 0 or not best_pattern:
        return []

    matches = list(best_pattern.finditer(text))
    out: list[dict] = []
    leading_text = text[:matches[0].start()]
    attach_leading_to_first = _looks_like_leading_reading_material(leading_text)
    for i, m in enumerate(matches):
        start = 0 if i == 0 and attach_leading_to_first else m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        seg_text = text[start:end].strip()
        if not seg_text:
            continue
        qno_match = re.search(r"(\d+)", m.group(0))
        out.append({
            "text": seg_text,
            "source_question_no": qno_match.group(1) if qno_match else None,
        })
    return out


def _split_by_char_count(text: str, max_chars: int = CHUNK_MAX_CHARS) -> list[dict]:
    """按字符数粗切文本（当没有题号模式时的 fallback）"""
    if len(text) <= max_chars:
        return [{"text": text, "start_page": None, "end_page": None}]

    segments = []
    start = 0
    while start < len(text):
        end = start + max_chars
        if end < len(text):
            # 在最近的换行处断开
            newline_pos = text.rfind("\n", start + CHUNK_MIN_CHARS, end)
            if newline_pos > start:
                end = newline_pos
        segment_text = text[start:end].strip()
        if segment_text:
            segments.append({"text": segment_text, "start_page": None, "end_page": None})
        start = end
    return segments


# ─── LLM 交互 ──────────────────────────────────


def _build_llm_prompt(chunk_text: str, answer_key_text: str = "") -> list[dict]:
    """构建 LLM 请求 messages"""
    answer_key_section = ""
    if answer_key_text:
        answer_key_section = (
            "\n以下是本文档的答案参考表，请结合此表确定每道题的答案：\n"
            f"{answer_key_text}\n"
        )

    system_content = (
        "你是一个 PDF/DOCX/XLSX 题库结构化解析器。\n"
        "\n"
        "要求：\n"
        "1. 只根据原文抽取题目，不要编造。\n"
        "2. 一个文本片段可能包含一道题或多道题。\n"
        "3. 识别题干（content）、选项（options）、答案（correct_answer）、解析（explanation）、参考资料（references）。\n"
        "4. 如果是场景题或长阅读材料题，把完整案例背景/阅读材料放入 scenario 字段，最后一问放入 content 字段；系统会将 scenario 与 content 合并为正式题干。\n"
        "5. 不要遗漏题号前紧邻的 SCENARIO、Case Study、阅读材料；如果无法确定场景背景是否完整，请降低 confidence 并在 issues 说明。\n"
        "6. 答案必须来自原文中的 Answer / Correct Answer / Answer Key 等标记"
        + ("或答案参考表" if answer_key_text else "") + "。\n"
        "7. 如果原文没有答案，correct_answer 输出空数组。\n"
        "8. 忽略页眉、页脚、广告、水印、文件版本信息等无关内容。\n"
        "9. 输出必须是 JSON，不要输出 Markdown 代码块标记。\n"
        "10. question_type 取值: single（单选）、multiple（多选）、truefalse（判断题）、unknown（不确定）。\n"
        "11. 如果选项标记有 [CORRECT] 或类似标记，该选项即为正确答案。\n"
        "12. 只提取有明确题号标记的题目（如 Q1、Question #1、1. 等编号格式），"
        "忽略没有题号标记的段落、知识点讲解、案例分析说明等非题目内容。\n"
        "13. 如果一段文字是知识讲解而非考试题目，不要将其转化为题目格式。\n"
        f"{answer_key_section}"
        "\n"
        'JSON 格式要求：\n'
        '{\n'
        '  "questions": [\n'
        '    {\n'
        '      "source_question_no": "题号",\n'
        '      "question_type": "single",\n'
        '      "scenario": null,\n'
        '      "content": "题干文本",\n'
        '      "options": [\n'
        '        {"label": "A", "text": "选项A文本"},\n'
        '        {"label": "B", "text": "选项B文本"},\n'
        '        {"label": "C", "text": "选项C文本"},\n'
        '        {"label": "D", "text": "选项D文本"}\n'
        '      ],\n'
        '      "correct_answer": ["A"],\n'
        '      "explanation": "",\n'
        '      "references": [],\n'
        '      "confidence": 0.95,\n'
        '      "issues": []\n'
        '    }\n'
        '  ],\n'
        '  "chunk_issues": []\n'
        '}'
    )

    user_content = f"请解析以下题库文本片段：\n\n{chunk_text}"

    return [
        {"role": "system", "content": system_content},
        {"role": "user", "content": user_content},
    ]


def _source_text_for_quality_check(parsed_q: ParsedQuestion, chunk_text: str) -> str:
    """定位当前题目的原文片段，用于场景题质量检查。

    L1 解析时一个 chunk 可能包含多道题；不能因为 chunk 中某一道题有 SCENARIO，
    就把同 chunk 的普通题都判为缺失场景。能按题号定位时只检查当前题片段，
    定位失败时再回退到完整 chunk。
    """
    normalized_qno = _normalize_qno(parsed_q.source_question_no)
    if normalized_qno is None:
        return chunk_text or ""

    for segment in _split_by_single_question(chunk_text or ""):
        if _normalize_qno(segment.get("source_question_no")) == normalized_qno:
            return segment.get("text") or ""

    return chunk_text or ""


def _parse_llm_response(response_text: str) -> LlmParseResult:
    """解析 LLM 的 JSON 响应，使用 Pydantic 校验"""
    text = response_text.strip()

    # 去除可能的 Markdown 代码块标记
    if text.startswith("```"):
        lines = text.split("\n", 1)
        if len(lines) > 1:
            text = lines[1]
        text = text.rsplit("```", 1)[0]
        text = text.strip()

    # 尝试直接解析
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # 尝试提取 JSON 部分
        json_match = re.search(r'\{[\s\S]*\}', text)
        if json_match:
            data = json.loads(json_match.group())
        else:
            raise ValueError("LLM 响应不是合法的 JSON")

    # Pydantic 校验
    return LlmParseResult.model_validate(data)


# ─── 质量检查 ──────────────────────────────────


def _quality_check(
    parsed_q: ParsedQuestion,
    chunk_text: str,
) -> tuple[Decimal, list[dict]]:
    """质量评分，返回 (final_confidence, issues)

    评分公式：
      final_confidence = 0.25 * llm_confidence
                       + 0.20 * schema_score
                       + 0.20 * answer_evidence_score
                       + 0.15 * option_quality_score
                       + 0.10 * duplicate_safety_score
                       + 0.10 * noise_clean_score
    """
    issues = []

    # 1. schema_score: 基本字段完整性
    schema_score = 1.0
    full_content = build_full_question_content(parsed_q.scenario, parsed_q.content)
    if not full_content or len(full_content.strip()) < 10:
        schema_score *= 0.5
        issues.append({"code": "STEM_TOO_SHORT", "severity": "MEDIUM", "detail": "题干过短"})

    source_text = _source_text_for_quality_check(parsed_q, chunk_text)
    source_has_scenario_marker = re.search(r"\b(SCENARIO|CASE\s+STUDY)\b", source_text or "", re.IGNORECASE)
    content_has_scenario_marker = re.search(r"\b(SCENARIO|CASE\s+STUDY)\b", full_content or "", re.IGNORECASE)
    if source_has_scenario_marker and not parsed_q.scenario and not content_has_scenario_marker:
        schema_score *= 0.6
        issues.append({
            "code": "SCENARIO_MISSING",
            "severity": "HIGH",
            "detail": "原文疑似包含场景/阅读材料，但解析结果未保留场景背景",
        })
    if len(parsed_q.options) < 2:
        schema_score *= 0.3
        issues.append({"code": "OPTION_COUNT_ABNORMAL", "severity": "HIGH", "detail": f"选项数量异常: {len(parsed_q.options)}"})
    if not parsed_q.correct_answer:
        issues.append({"code": "ANSWER_MISSING", "severity": "HIGH", "detail": "未找到答案"})

    # 2. answer_evidence_score: 答案是否在选项中
    answer_evidence_score = 1.0
    if parsed_q.correct_answer:
        option_labels = {opt.label.upper() for opt in parsed_q.options}
        for ans in parsed_q.correct_answer:
            if ans.upper() not in option_labels:
                answer_evidence_score *= 0.3
                issues.append({
                    "code": "ANSWER_NOT_IN_OPTIONS",
                    "severity": "HIGH",
                    "detail": f"答案 {ans} 不在选项中",
                })
                break
    else:
        answer_evidence_score = 0.0

    # 3. option_quality_score: 选项数量和质量
    option_quality_score = 1.0
    opt_count = len(parsed_q.options)
    if opt_count < 2 or opt_count > 8:
        option_quality_score = 0.5
    elif opt_count < 4:
        option_quality_score = 0.8
    # 检查选项文本是否过短
    short_options = sum(1 for opt in parsed_q.options if len(opt.text.strip()) < 3)
    if short_options > 0:
        option_quality_score *= 0.8

    # 4. duplicate_safety_score: 无题号标记的题目降分
    if not parsed_q.source_question_no or parsed_q.source_question_no.strip().lower() in ("unknown", ""):
        duplicate_safety_score = 0.5
        issues.append({"code": "NO_QUESTION_NO", "severity": "MEDIUM", "detail": "无题号标记，可能不是正式题目"})
    else:
        duplicate_safety_score = 1.0

    # 5. noise_clean_score: 噪声检测
    noise_clean_score = 1.0
    combined_text = full_content + " " + " ".join(opt.text for opt in parsed_q.options)
    for noise_pattern in NOISE_PATTERNS:
        if noise_pattern.search(combined_text):
            noise_clean_score *= 0.7
            issues.append({
                "code": "NOISE_DETECTED",
                "severity": "LOW",
                "detail": f"检测到可能的噪声内容",
            })
            break

    # 计算最终置信度
    llm_conf = float(parsed_q.confidence)
    final_confidence = (
        0.25 * llm_conf
        + 0.20 * schema_score
        + 0.20 * answer_evidence_score
        + 0.15 * option_quality_score
        + 0.10 * duplicate_safety_score
        + 0.10 * noise_clean_score
    )
    final_confidence = round(final_confidence, 4)

    return Decimal(str(final_confidence)), issues


# ─── 缓存管理 ──────────────────────────────────


def _build_cache_key(chunk_hash: str) -> str:
    """构建缓存键: sha256(model_name + prompt_version + chunk_hash)"""
    raw = f"{PROMPT_VERSION}:{chunk_hash}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _lookup_llm_cache(db: Session, cache_key: str) -> dict | None:
    """查找 LLM 缓存"""
    cached = db.query(LlmParseCache).filter_by(cache_key=cache_key).first()
    if not cached:
        return None
    return {
        "response_text": json.dumps(cached.response_json, ensure_ascii=False) if cached.response_json else "",
        "request_json": cached.request_json,
    }


def _store_llm_cache(
    db: Session,
    cache_key: str,
    chunk_hash: str,
    request_json: dict | None = None,
    response_text: str = "",
) -> None:
    """存储 LLM 缓存"""
    response_json = None
    if response_text:
        try:
            response_json = json.loads(response_text)
        except json.JSONDecodeError:
            response_json = {"raw": response_text[:2000]}

    cache_entry = LlmParseCache(
        cache_key=cache_key,
        model_name=None,
        prompt_version=PROMPT_VERSION,
        chunk_hash=chunk_hash,
        request_json=request_json,
        response_json=response_json,
    )
    db.add(cache_entry)
    db.flush()


# ─── ImportJob 状态管理 ──────────────────────────────────


IMPORT_JOB_STATUSES = {
    "pending", "extracting", "chunking", "parsing",
    "validating", "importing", "imported",
    "review_required", "unimported", "partial_imported", "failed", "cancelled",
}


def _update_import_job_status(db: Session, import_job: ImportJob, status: str) -> None:
    """更新 ImportJob 状态"""
    import_job = db.get(ImportJob, import_job.id)
    if import_job:
        import_job.status = status
        import_job.updated_at = datetime.now(timezone.utc)
        db.commit()


def _fail_import_job(db: Session, import_job: ImportJob, error_message: str) -> None:
    """标记 ImportJob 为失败"""
    import_job = db.get(ImportJob, import_job.id)
    if import_job:
        import_job.status = "failed"
        import_job.error_message = error_message
        import_job.updated_at = datetime.now(timezone.utc)
        db.commit()


def _compute_reconciliation(db: Session, import_job: ImportJob) -> dict:
    """汇总 expected / imported_unique / missing_qnos / duplicates_in_db。

    数据来源（PR-2/PR-3 已写入字段）：
      - expected_qnos       : import_job.config_json["expected_qnos"]
                              （run_smart_import 切完 chunk 后写入；reparse 不变）
      - imported_unique     : ImportParsedQuestion.source_question_no
                              WHERE import_status='imported'
      - duplicates_in_db    : ImportParsedQuestion.source_question_no
                              WHERE import_status='skipped'
                              AND issues_json.details[0].reason == 'qno' (PR-3 接口)
      - per_question_failures :
                              ImportChunk.issues_json["per_question_failures"][*].source_question_no
                              （PR-2 L2 fallback 写入）

    missing_qnos = (expected - imported_unique) ∪ per_question_failures
    （合集语义保护"LLM 输出题号格式漂移让 expected_qnos 漏题号"的极端情况）
    """
    config = import_job.config_json or {}
    expected_set: set[str] = set(config.get("expected_qnos") or [])

    imported_set: set[str] = set()
    for pq in (
        db.query(ImportParsedQuestion)
        .filter_by(import_job_id=import_job.id, import_status="imported")
        .all()
    ):
        norm = _normalize_qno(pq.source_question_no)
        if norm is not None:
            imported_set.add(norm)

    duplicates_set: set[str] = set()
    for pq in (
        db.query(ImportParsedQuestion)
        .filter_by(import_job_id=import_job.id, import_status="skipped")
        .all()
    ):
        details = (pq.issues_json or {}).get("details") or []
        if any(d.get("reason") == "qno" for d in details):
            norm = _normalize_qno(pq.source_question_no)
            if norm is not None:
                duplicates_set.add(norm)

    per_q_failures_set: set[str] = set()
    for chunk in (
        db.query(ImportChunk).filter_by(import_job_id=import_job.id).all()
    ):
        failures = (chunk.issues_json or {}).get("per_question_failures") or []
        for f in failures:
            norm = _normalize_qno(f.get("source_question_no"))
            if norm is not None:
                per_q_failures_set.add(norm)

    missing_set = (expected_set - imported_set) | per_q_failures_set

    return {
        "expected": sorted(expected_set, key=_qno_sort_key),
        "imported_unique": sorted(imported_set, key=_qno_sort_key),
        "missing_qnos": sorted(missing_set, key=_qno_sort_key),
        "duplicates_in_db": sorted(duplicates_set, key=_qno_sort_key),
        "per_question_failures_count": len(per_q_failures_set),
        "computed_at": datetime.now(timezone.utc).isoformat(),
    }


def _write_reconciliation(db: Session, import_job: ImportJob) -> None:
    """计算 reconciliation 报告并写入 import_job.config_json。

    用 dict 重赋值触发 SQLAlchemy JSONB dirty 检测；不依赖 flag_modified。
    与 run_smart_import / run_reparse 内现有 config_json 写入姿势一致。
    """
    recon = _compute_reconciliation(db, import_job)
    import_job.config_json = {
        **(import_job.config_json or {}),
        "reconciliation": recon,
    }
    db.commit()


def _auto_handled_counts(db: Session, import_job_id: int) -> dict:
    auto_imported = db.query(func.count(ImportParsedQuestion.id)).filter_by(
        import_job_id=import_job_id,
        review_status="auto_accepted",
        import_status="imported",
    ).scalar() or 0
    auto_skipped = db.query(func.count(ImportParsedQuestion.id)).filter_by(
        import_job_id=import_job_id,
        review_status="auto_skipped",
        import_status="skipped",
    ).scalar() or 0
    return {
        "auto_imported": auto_imported,
        "auto_skipped": auto_skipped,
        "auto_handled": auto_imported + auto_skipped,
    }


def _finalize_import(db: Session, import_job: ImportJob) -> None:
    """导入完成后的汇总处理"""
    import_job = db.get(ImportJob, import_job.id)
    if not import_job:
        return

    # 统计最终结果
    total_parsed = import_job.parsed_questions or 0
    total_imported = import_job.imported_questions or 0
    total_review = import_job.review_questions or 0
    total_failed = import_job.failed_chunks or 0
    auto_counts = _auto_handled_counts(db, import_job.id)
    total_auto_skipped = auto_counts["auto_skipped"]

    # 生成摘要
    summary = {
        "total_parsed": total_parsed,
        "total_imported": total_imported,
        "total_review": total_review,
        "total_failed": total_failed,
        **auto_counts,
        "auto_import_rate": round(total_imported / total_parsed, 4) if total_parsed > 0 else 0,
    }
    import_job.summary_json = summary

    # 确定最终状态
    if total_review > 0 and total_imported > 0:
        import_job.status = "partial_imported"
    elif total_review > 0:
        import_job.status = "review_required"
    elif total_failed > 0:
        import_job.status = "partial_imported"
    elif total_imported > 0 and total_auto_skipped > 0:
        import_job.status = "partial_imported"
    elif total_imported > 0:
        import_job.status = "imported"
    elif total_auto_skipped > 0 and total_parsed > 0:
        import_job.status = "unimported"
    else:
        import_job.status = "imported"

    import_job.updated_at = datetime.now(timezone.utc)
    db.commit()

    # 更新题库统计和高频词
    if total_imported > 0:
        _update_bank_stats(db, import_job.bank_id)

    # 更新题库源文件名
    bank = db.get(QuestionBank, import_job.bank_id)
    if bank:
        bank.source_filename = import_job.file_name
        db.commit()

    # PR-4：写入 reconciliation 报告
    # 放在 status / summary_json / bank stats 落盘之后；reconciliation 是独立的"对账"
    # 元数据，不影响 ImportJob.status 的状态机决策。
    _write_reconciliation(db, import_job)


def _update_bank_stats(db: Session, bank_id: int) -> None:
    """更新题库题目数量和高频词统计"""
    bank = db.get(QuestionBank, bank_id)
    if not bank:
        return

    # 更新题目数量
    total_questions = db.query(func.count(Question.id)).filter_by(bank_id=bank_id).scalar() or 0
    bank.question_count = total_questions

    # 重建高频词统计（保留已有翻译）
    existing_translations = {
        row.term: row.term_zh
        for row in db.query(BankWordFrequency).filter_by(bank_id=bank_id).all()
        if row.term_zh
    }

    all_questions = (
        db.query(Question)
        .filter_by(bank_id=bank_id)
        .order_by(Question.order_index.asc(), Question.id.asc())
        .all()
    )

    frequency_items = build_bank_word_frequencies([
        {
            "content": q.content,
            "options": q.options if isinstance(q.options, list) else json.loads(q.options),
        }
        for q in all_questions
    ])

    excluded_terms = {
        row.term
        for row in db.query(BankWordExclusion).filter_by(bank_id=bank_id).all()
    }

    db.query(BankWordFrequency).filter_by(bank_id=bank_id).delete()
    for item in frequency_items:
        if item["term"] in excluded_terms:
            continue
        db.add(BankWordFrequency(
            bank_id=bank_id,
            term=item["term"],
            term_zh=existing_translations.get(item["term"]),
            frequency=item["frequency"],
        ))

    db.commit()


# ─── 序列化辅助 ──────────────────────────────────


def serialize_import_job(import_job: ImportJob, db: Session | None = None) -> dict:
    """将 ImportJob 序列化为字典"""
    bg_job = None
    if import_job.background_job_id:
        bg_job_record = import_job.background_job
        if bg_job_record:
            bg_job = {
                "id": bg_job_record.id,
                "status": bg_job_record.status,
                "progress_total": bg_job_record.progress_total,
                "progress_done": bg_job_record.progress_done,
                "status_message": bg_job_record.status_message,
            }

    summary = import_job.summary_json or {}
    auto_counts = _auto_handled_counts(db, import_job.id) if db else {
        "auto_imported": summary.get("auto_imported", 0),
        "auto_skipped": summary.get("auto_skipped", 0),
        "auto_handled": summary.get("auto_handled", 0),
    }

    return {
        "id": import_job.id,
        "bank_id": import_job.bank_id,
        "background_job_id": import_job.background_job_id,
        "file_name": import_job.file_name,
        "file_type": import_job.file_type,
        "status": import_job.status,
        "total_pages": import_job.total_pages,
        "total_chunks": import_job.total_chunks,
        "parsed_questions": import_job.parsed_questions,
        "imported_questions": import_job.imported_questions,
        "review_questions": import_job.review_questions,
        "failed_chunks": import_job.failed_chunks,
        "auto_imported_questions": auto_counts["auto_imported"],
        "auto_skipped_questions": auto_counts["auto_skipped"],
        "auto_handled_questions": auto_counts["auto_handled"],
        "summary": import_job.summary_json,
        "error_message": import_job.error_message,
        "created_by": import_job.created_by,
        "created_at": import_job.created_at.isoformat() if import_job.created_at else None,
        "updated_at": import_job.updated_at.isoformat() if import_job.updated_at else None,
        "background_job": bg_job,
        "reconciliation": (import_job.config_json or {}).get("reconciliation"),
    }


def serialize_chunk(chunk: ImportChunk) -> dict:
    """将 ImportChunk 序列化为字典"""
    return {
        "id": chunk.id,
        "import_job_id": chunk.import_job_id,
        "chunk_no": chunk.chunk_no,
        "start_page": chunk.start_page,
        "end_page": chunk.end_page,
        "chunk_text": chunk.chunk_text,
        "status": chunk.status,
        "issues": chunk.issues_json,
        "created_at": chunk.created_at.isoformat() if chunk.created_at else None,
    }


def serialize_parsed_question(pq: ImportParsedQuestion) -> dict:
    """将 ImportParsedQuestion 序列化为字典"""
    correct_answer = pq.correct_answer
    if isinstance(correct_answer, str):
        correct_answer = [a.strip() for a in correct_answer.split(",") if a.strip()]

    return {
        "id": pq.id,
        "import_job_id": pq.import_job_id,
        "chunk_id": pq.chunk_id,
        "source_question_no": pq.source_question_no,
        "question_type": pq.question_type,
        "scenario_text": pq.scenario_text,
        "content": pq.content,
        "options": pq.options_json,
        "correct_answer": correct_answer,
        "explanation": pq.explanation,
        "llm_confidence": float(pq.llm_confidence) if pq.llm_confidence else None,
        "final_confidence": float(pq.final_confidence) if pq.final_confidence else None,
        "issues": pq.issues_json,
        "review_status": pq.review_status,
        "import_status": pq.import_status,
        "imported_question_id": pq.imported_question_id,
    }


UNUSABLE_REASON_LABELS = {
    "STEM_MISSING": "缺少题干",
    "STEM_TOO_SHORT": "题干过短",
    "OPTIONS_MISSING": "选项不足",
    "OPTION_COUNT_ABNORMAL": "选项数量异常",
    "ANSWER_MISSING": "缺少正确答案",
    "ANSWER_NOT_IN_OPTIONS": "正确答案不在选项中",
}

QUALITY_TIP_LABELS = {
    "LOW_CONFIDENCE": "置信度较低",
    "NO_QUESTION_NO": "无题号",
    "NOISE_DETECTED": "疑似包含噪声",
    "SCENARIO_MISSING": "疑似缺少场景材料",
    "OPTION_COUNT_ABNORMAL": "选项数量异常",
    "STEM_TOO_SHORT": "题干偏短",
}


def _issue_details(pq: ImportParsedQuestion) -> list[dict]:
    issues = pq.issues_json or {}
    details = issues.get("details") if isinstance(issues, dict) else None
    return details if isinstance(details, list) else []


def _auto_handled_reason(pq: ImportParsedQuestion) -> str:
    if pq.review_status == "auto_skipped":
        for issue in _issue_details(pq):
            code = issue.get("code")
            if code in UNUSABLE_REASON_LABELS:
                return UNUSABLE_REASON_LABELS[code]
        return "不可用题目，已自动跳过"
    return "题目结构完整，已自动入库"


def _auto_handled_quality_tips(pq: ImportParsedQuestion) -> list[str]:
    tips = []
    for issue in _issue_details(pq):
        code = issue.get("code")
        if pq.review_status == "auto_skipped" and code in UNUSABLE_REASON_LABELS:
            continue
        label = QUALITY_TIP_LABELS.get(code)
        if label and label not in tips:
            tips.append(label)
    return tips


def serialize_auto_handled_item(pq: ImportParsedQuestion) -> dict:
    return {
        "id": pq.id,
        "import_job_id": pq.import_job_id,
        "result": "auto_skipped" if pq.review_status == "auto_skipped" else "auto_imported",
        "reason": _auto_handled_reason(pq),
        "quality_tips": _auto_handled_quality_tips(pq),
        "source_question_no": pq.source_question_no,
        "content": pq.content,
        "scenario_text": pq.scenario_text,
        "question_type": pq.question_type,
        "options": pq.options_json,
        "correct_answer": pq.correct_answer,
        "imported_question_id": pq.imported_question_id,
        "handled_at": pq.updated_at.isoformat() if pq.updated_at else (pq.created_at.isoformat() if pq.created_at else None),
        "issues": pq.issues_json,
    }


def serialize_review_item(ri: ImportReviewItem, parsed_question: ImportParsedQuestion | None = None, db: Session | None = None) -> dict:
    """将 ImportReviewItem 序列化为字典，含关联的解析题目和原始 chunk"""
    result = {
        "id": ri.id,
        "import_job_id": ri.import_job_id,
        "parsed_question_id": ri.parsed_question_id,
        "review_type": ri.review_type,
        "severity": ri.severity,
        "before_json": ri.before_json,
        "after_json": ri.after_json,
        "status": ri.status,
        "reviewer_id": ri.reviewer_id,
        "reviewed_at": ri.reviewed_at.isoformat() if ri.reviewed_at else None,
        "parsed_question": None,
        "chunk_text": None,
    }

    if parsed_question:
        result["parsed_question"] = serialize_parsed_question(parsed_question)
        # 获取关联的 chunk 文本
        if parsed_question.chunk_id and db:
            chunk = db.get(ImportChunk, parsed_question.chunk_id)
            if chunk:
                result["chunk_text"] = chunk.chunk_text

    return result


# ─── 内部辅助 ──────────────────────────────────


def _deserialize_payload(job: BackgroundJob) -> dict:
    """反序列化 BackgroundJob 的 payload_json"""
    if not job.payload_json:
        return {}
    try:
        return json.loads(job.payload_json)
    except json.JSONDecodeError:
        return {}
