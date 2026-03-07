import re
import io
import json
from collections import Counter


MIN_FREQUENCY = 2
TOP_FREQUENT_TERMS_LIMIT = 5000
WORD_PATTERN = re.compile(r"[A-Za-z]+(?:'[A-Za-z]+)?")
EXCLUDED_TERMS = {
    'a', 'an', 'the',
    'i', 'me', 'my', 'mine', 'myself', 'we', 'us', 'our', 'ours', 'ourselves',
    'you', 'your', 'yours', 'yourself', 'yourselves', 'he', 'him', 'his', 'himself',
    'she', 'her', 'hers', 'herself', 'it', 'its', 'itself', 'they', 'them', 'their',
    'theirs', 'themselves', 'this', 'that', 'these', 'those', 'who', 'whom', 'whose',
    'which', 'what', 'whatever', 'whoever', 'whichever', 'someone', 'somebody', 'something',
    'anyone', 'anybody', 'anything', 'everyone', 'everybody', 'everything', 'none', 'nothing',
    'one', 'ones', 'another', 'others', 'other', 'each', 'either', 'neither', 'both', 'all',
    'few', 'many', 'several', 'some', 'any', 'most', 'such',
    'zero', 'first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh', 'eighth', 'ninth',
    'tenth', 'eleventh', 'twelfth', 'thirteenth', 'fourteenth', 'fifteenth', 'sixteenth',
    'seventeenth', 'eighteenth', 'nineteenth', 'twentieth', 'once', 'twice',
    'in', 'on', 'at', 'by', 'for', 'from', 'to', 'of', 'with', 'without', 'within', 'into', 'onto',
    'over', 'under', 'above', 'below', 'between', 'among', 'through', 'during', 'before', 'after',
    'since', 'until', 'upon', 'about', 'across', 'against', 'around', 'behind', 'beyond', 'beside',
    'besides', 'near', 'inside', 'outside', 'toward', 'towards', 'via', 'per', 'past',
    'and', 'or', 'but', 'nor', 'so', 'yet', 'for', 'although', 'though', 'because', 'if', 'unless',
    'while', 'whereas', 'whether', 'than', 'once',
    'am', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
    'do', 'does', 'did', 'done', 'doing', 'have', 'has', 'had', 'having',
    'can', 'could', 'may', 'might', 'must', 'shall', 'should', 'will', 'would',
    'not', 'no', 'yes', 'as',
    'oh', 'ah', 'wow', 'hey', 'oops', 'alas', 'hurray', 'bravo'
}


def parse_file(file_storage, filename):
    content = ''
    if filename.endswith('.pdf'):
        content = _extract_pdf(file_storage)
    elif filename.endswith('.xlsx'):
        content = _extract_xlsx(file_storage)
    elif filename.endswith('.docx'):
        content = _extract_docx(file_storage)
    else:
        raise ValueError(f'不支持的文件格式: {filename}')

    return _parse_questions(content)


def _clean_text(text):
    """清理 PDF 提取文本中的控制字符和常见 ligature 问题"""
    # PDF 中 ligature（fi, fl, ff 等）有时提取为 \x00
    replacements = {
        '\x00rst': 'first', '\x00nd': 'find', '\x00le': 'file',
        '\x00re': 'fire', '\x00x': 'fix', '\x00eld': 'field',
        '\x00lter': 'filter', '\x00rm': 'firm', '\x00nal': 'final',
        '\x00nancial': 'financial', '\x00gure': 'figure',
        '\x00ll': 'fill', '\x00nger': 'finger',
        '\x00ber': 'fiber', '\x00ve': 'five',
        'certi\x00cate': 'certificate', 'certi\x00ed': 'certified',
        'con\x00dential': 'confidential', 'con\x00gur': 'configur',
        'con\x00rm': 'confirm', 'con\x00ned': 'confined',
        'identi\x00': 'identifi', 'speci\x00': 'specifi',
        'signi\x00': 'signifi', 'noti\x00': 'notifi',
        'bene\x00t': 'benefit', 'pro\x00le': 'profile',
        'Paci\x00c': 'Pacific', 'e\x00ect': 'effect',
        'o\x00ce': 'office', 'o\x00er': 'offer',
        'o\x00ine': 'offline', 'o\x00icial': 'official',
        'su\x00cient': 'sufficient', 'e\x00cien': 'efficien',
        'di\x00cult': 'difficult', 'di\x00erent': 'different',
        'a\x00ect': 'affect', 'a\x00liat': 'affiliat',
        '\x00exib': 'flexib', '\x00ow': 'flow',
        '\x00ag': 'flag', '\x00at': 'flat',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    # 移除剩余的空字节和控制字符（保留换行和制表符）
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    return text


def _extract_pdf(file_storage):
    import pdfplumber
    text_parts = []
    with pdfplumber.open(file_storage) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
    return _clean_text('\n'.join(text_parts))


def _extract_xlsx(file_storage):
    from openpyxl import load_workbook
    wb = load_workbook(file_storage, read_only=True)
    lines = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            line = ' '.join(str(c) for c in row if c is not None)
            if line.strip():
                lines.append(line)
    return '\n'.join(lines)


def _extract_docx(file_storage):
    from docx import Document
    doc = Document(file_storage)
    return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())


def build_bank_word_frequencies(questions, min_frequency=MIN_FREQUENCY):
    counter = Counter()

    for question in questions:
        counter.update(_extract_terms_from_question(question))

    items = []
    for term, frequency in counter.items():
        if frequency >= min_frequency:
            items.append({'term': term, 'frequency': frequency})

    return sorted(items, key=lambda item: (-item['frequency'], item['term']))


def _extract_terms_from_question(question):
    texts = [question.get('content', '')]
    texts.extend(option.get('text', '') for option in question.get('options', []))

    terms = []
    for text in texts:
        for token in WORD_PATTERN.findall(text or ''):
            normalized = _normalize_term(token)
            if normalized and not _should_exclude_term(normalized):
                terms.append(normalized)
    return terms


def _normalize_term(token):
    term = token.strip("'\".,;:!?()[]{}<>").lower()
    if term.endswith("'s"):
        term = term[:-2]
    elif term.endswith("s'"):
        term = term[:-1]
    return term.strip("'\"")


def _should_exclude_term(term):
    if not term:
        return True
    if len(term) <= 4:
        return True
    if term in EXCLUDED_TERMS:
        return True
    if re.fullmatch(r'\d+(?:st|nd|rd|th)?', term):
        return True
    return False


def _parse_questions(text):
    # 优先尝试考试题库格式（Question #N Topic N）
    exam_pattern = r'Question\s+#(\d+)\s+Topic\s+\d+'
    exam_splits = re.split(exam_pattern, text)
    if len(exam_splits) >= 3:
        return _parse_exam_dump(exam_splits)

    return _parse_generic(text)


def _parse_exam_dump(splits):
    """解析考试题库格式（如 ExamTopics 导出的 PDF）"""
    questions = []

    for i in range(1, len(splits) - 1, 2):
        q_num = int(splits[i])
        q_text = splits[i + 1].strip()

        # 提取正确答案并截断后面的评论区
        answer = ''
        answer_match = re.search(
            r'Correct\s+Answer:\s*([A-E](?:\s*,\s*[A-E])*)',
            q_text, re.IGNORECASE
        )
        if answer_match:
            answer = answer_match.group(1).strip().upper().replace(' ', '')
            q_text = q_text[:answer_match.start()].strip()

        # 提取选项
        option_pattern = r'(?:^|\n)\s*([A-E])\.\s+(.*?)(?=\n\s*[A-E]\.\s|\n\s*Correct\s+Answer|$)'
        option_matches = re.findall(option_pattern, q_text, re.DOTALL)

        options = []
        for key, opt_text in option_matches:
            options.append({
                'key': key.upper(),
                'text': opt_text.strip().replace('\n', ' '),
            })

        # 提取题干
        if options:
            first_opt = re.search(r'\n\s*[A-E]\.\s+', q_text)
            stem = q_text[:first_opt.start()].strip() if first_opt else q_text
        else:
            stem = q_text

        # 清理 SCENARIO 标记，保留内容
        stem = re.sub(r'^SCENARIO\s*-?\s*\n?', '', stem).strip()
        stem = re.sub(r'\n+', '\n', stem).strip()

        if not stem or not options:
            continue

        q_type = 'multiple' if ',' in answer else 'single'

        questions.append({
            'content': stem,
            'options': options,
            'correct_answer': answer or '',
            'question_type': q_type,
            'answer_missing': not answer,
        })

    return questions


def _parse_generic(text):
    """解析通用题目格式（如 1. / Q1. / Question 1:）"""
    questions = []

    # 尝试在末尾找答案表
    answer_key = {}
    answer_key_match = re.search(
        r'(?:Answer\s*Key|Answers?|ANSWER\s*KEY|正确答案)[:\s]*\n([\s\S]+?)$',
        text, re.IGNORECASE
    )
    if answer_key_match:
        key_text = answer_key_match.group(1)
        for m in re.finditer(r'(\d+)[.\s)]+\s*([A-Ea-e](?:\s*[,，]\s*[A-Ea-e])*|True|False)', key_text):
            answer_key[int(m.group(1))] = m.group(2).strip().upper().replace('，', ',')
        text = text[:answer_key_match.start()]

    # 按题号分割
    pattern = r'(?:^|\n)\s*(?:Q(?:uestion)?\s*)?(\d+)\s*[.)\]:]\s*'
    splits = re.split(pattern, text)

    for i in range(1, len(splits) - 1, 2):
        q_num = int(splits[i])
        q_text = splits[i + 1].strip()

        # 提取选项
        option_pattern = r'(?:^|\n)\s*([A-Ea-e])\s*[.)]\s*(.*?)(?=(?:\n\s*[A-Ea-e]\s*[.)]|\n\s*(?:Answer|Correct|正确)|$))'
        option_matches = re.findall(option_pattern, q_text, re.DOTALL | re.IGNORECASE)

        options = []
        for key, opt_text in option_matches:
            options.append({
                'key': key.upper(),
                'text': opt_text.strip(),
            })

        if options:
            first_opt_pattern = r'\n?\s*[A-Ea-e]\s*[.)]\s*'
            stem = re.split(first_opt_pattern, q_text)[0].strip()
        else:
            stem = q_text

        if not stem:
            continue

        answer = ''
        answer_match = re.search(
            r'(?:Answer|Correct(?:\s*Answer)?)\s*[:\s]+\s*([A-Ea-e](?:\s*[,，]\s*[A-Ea-e])*|True|False)',
            q_text, re.IGNORECASE
        )
        if answer_match:
            answer = answer_match.group(1).strip().upper().replace('，', ',')

        if not answer and q_num in answer_key:
            answer = answer_key[q_num]

        if answer in ('TRUE', 'FALSE'):
            q_type = 'truefalse'
        elif ',' in answer:
            q_type = 'multiple'
        else:
            q_type = 'single'

        if options or q_type == 'truefalse':
            if q_type == 'truefalse' and not options:
                options = [
                    {'key': 'A', 'text': 'True'},
                    {'key': 'B', 'text': 'False'},
                ]
                if answer == 'TRUE':
                    answer = 'A'
                elif answer == 'FALSE':
                    answer = 'B'

            questions.append({
                'content': stem,
                'options': options,
                'correct_answer': answer or '',
                'question_type': q_type,
                'answer_missing': not answer,
            })

    return questions
